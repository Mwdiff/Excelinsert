from DBcm import UseDatabase
from openpyxl import load_workbook
from init import dbconfig


def AppendSelectColumns(sheet: 'worksheet', columns: str, start_row: int=1, index_column: int=1) -> None:
    
    max_column = sheet.max_column # Data boundary in Excel sheet

    #Column names
    for (column_name, column_number) in zip(columns.split(', '), range(max_column+1, 100)):
        sheet.cell(start_row, column_number).value = column_name

    with UseDatabase(dbconfig) as cursor:
        for cell in sheet.iter_rows(min_row=start_row+1, max_row=sheet.max_row, min_col=index_column, max_col=index_column):
            _SQL = f'select {columns} from towary where symbol=\'{str(cell[0].value).zfill(6)}\''
            print(_SQL)
            cursor.execute(_SQL)

            try:
                data = cursor.fetchall()[0]
            except IndexError:
                data = ['NULL']*len(columns.split(', '))

            print(data)
            for i in range(len(data)):
                sheet.cell(cell[0].row, max_column+1+i).value = data[i]

# open file
file = input('Podaj nazwę pliku: ')

wb = load_workbook(f'./{file}.xlsx')
ws = wb[wb.sheetnames[0]]
last_column = ws.max_column
# last_row = ws.max_row
# print(ws['A2:A'+str(ws.max_row)])
# for cell in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=1, values_only=1):
#     print(cell[0])

AppendSelectColumns(ws, 'detal, zakup', start_row = 4, index_column = 1)
wb.save(f'./edit_{file}.xlsx')